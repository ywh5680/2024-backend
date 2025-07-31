
import csv
from typing import Callable
from django.http import HttpResponse
from django.db.models import Model
from django.contrib.admin import ModelAdmin
from . import models
from . import serializers
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def export_csv(self: ModelAdmin, request, queryset) -> HttpResponse:
    Ser = serializers.EnrollSerializer

    meta = self.model._meta
    field_names = list(field.name for field in meta.fields)
    try: field_names.remove("id")
    except ValueError: pass
    getfield = lambda obj, field: getattr(obj, field)
    if type(queryset.first()) is models.EnrollModel:
        getfield = lambda obj, field: Ser(obj).data.get(field, getattr(obj, field))
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename={}.csv'.format(meta)
    writer = csv.writer(response)

    writer.writerow(field_names)
    for obj in queryset:
        writer.writerow(
            getfield(obj, field) for field in field_names
        )

    return response

export_csv.short_description = "导出所选为CSV"

def export_excel(self: ModelAdmin, request, queryset) -> HttpResponse:
    """导出Excel文件"""
    Ser = serializers.EnrollSerializer
    
    meta = self.model._meta
    field_names = list(field.name for field in meta.fields)
    try: field_names.remove("id")
    except ValueError: pass
    
    getfield = lambda obj, field: getattr(obj, field)
    if type(queryset.first()) is models.EnrollModel:
        getfield = lambda obj, field: Ser(obj).data.get(field, getattr(obj, field))
    
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "数据导出"
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # 写入表头
    for col, field in enumerate(field_names, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = header_font
        cell.fill = header_fill
    
    # 写入数据
    for row, obj in enumerate(queryset, 2):
        for col, field in enumerate(field_names, 1):
            ws.cell(row=row, column=col, value=str(getfield(obj, field)))
    
    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # 创建响应
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'
    
    # 保存文件
    wb.save(response)
    return response

export_excel.short_description = "导出所选为Excel"

class ExportCsvMixin:
    '''用于被AdminModel继承'''
    model: Model
    export_csv: Callable
    export_excel: Callable
ExportCsvMixin.export_csv = export_csv
ExportCsvMixin.export_excel = export_excel

